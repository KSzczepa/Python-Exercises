from openpyxl import load_workbook
print("START")

wb = load_workbook('sample.xlsx')
print(wb.sheetnames)
sheet_rangers = wb['Sheet']
print(sheet_rangers['A1'].value, sheet_rangers['B2'].value)
print(sheet_rangers['C3'].value)
print(sheet_rangers['D4'].value, sheet_rangers['E5'].value)