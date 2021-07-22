from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = 5
ws["A2"] = "Hello"
ws["B2"] = "People"
ws["C2"] = "from"
ws["D2"] = "PythonPlanet"
wb.save("Sample20.xlsx")
